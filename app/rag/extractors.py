"""Extracción de texto para formatos binarios de `docs/references/` (PDF, DOCX, XLSX).

Los `.md`/`.txt` no pasan por este módulo (siguen decodificándose directo en
`app/rag/ingestion.py::load_document`); este módulo cubre los formatos agregados
para poder indexar el corpus real de normativa/estándares de auditoría en
`docs/references/` (COBIT5, ISO 27001/27002, MAGERIT, guías ISACA, cuestionarios,
informes de ejemplo, herramientas CAAT).

SEGURIDAD (spec-005, `.ai/skills/security-prompt-injection/SKILL.md`): igual que en
`ingestion.py`/`chunking.py`, el texto que devuelven estas funciones es DATO
extraído de un documento ingerido, nunca una instrucción — estas funciones no
interpretan ni ejecutan nada del contenido, solo lo extraen como texto plano.

Tolerancia a fallos parciales (ver `app/rag/ingestion.py::_extract_units`):
- PDF: si una página individual falla al extraer texto (escaneada como imagen sin
  texto embebido, stream corrupto, etc.), se registra un `logger.warning`/`info` y
  esa página se devuelve con texto vacío en vez de abortar el documento entero
  (pypdf no hace OCR: una página escaneada como imagen produce texto vacío de forma
  esperada, no es un error). Si el PDF completo no se puede abrir (encriptado sin
  password que pypdf pueda descifrar, corrupto a nivel de contenedor), se levanta
  `ExtractionError` — ese archivo puntual falla, pero quien orquesta la ingesta en
  lote (`ingest_directory_recursive`) sigue con el resto de los archivos.
- DOCX/XLSX: si el archivo no se puede abrir, se levanta `ExtractionError`.

Fallback OCR por página (PDF únicamente): algunas copias de biblioteca del
corpus (`docs/references/Estandares/ISO 27001 ...pdf`, `ISO- 27002.pdf`) tienen
marca de agua/DRM y su contenido real está renderizado como imagen — `pypdf`
sólo recupera el texto de la marca de agua ("Licenciado por IRAM..."), no el
estándar en sí. Cuando el texto que extrae `pypdf` de una página individual
queda por debajo de `OCR_MIN_CHARS`, se renderiza esa página puntual a imagen
con PyMuPDF (`fitz`) y se le corre OCR con Tesseract (`pytesseract`, idioma
español). Es un fallback condicional por página, no un reemplazo del camino
rápido: para el resto del corpus (que ya extrae texto real con `pypdf`) el
umbral no debería dispararse y el costo de OCR no se paga. Si PyMuPDF o
Tesseract no están instalados, o el OCR también falla/da poco texto, se
conserva el texto original de `pypdf` para esa página — nunca se aborta el
documento completo por esto.
"""

from __future__ import annotations

import logging
import os
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)

# Debajo de esta cantidad de caracteres (tras `.strip()`), el texto que extrajo
# `pypdf` de una página se considera "posiblemente insuficiente" y se intenta el
# fallback OCR. Elegido para que las páginas reales de marca de agua/DRM del
# corpus (~220 caracteres, una sola línea de "Licenciado por IRAM...") caigan
# debajo del umbral, mientras que una página con contenido real de un estándar
# (varios párrafos) quede holgadamente por encima y no dispare OCR innecesario.
OCR_MIN_CHARS = 300

# Factor de zoom para renderizar la página a imagen antes del OCR. PyMuPDF
# renderiza a 72 DPI por defecto; un zoom de 3x da ~216 DPI, suficiente calidad
# para que Tesseract lea texto de cuerpo normal sin generar imágenes gigantes
# (que harían el OCR mucho más lento en los ~176 páginas del corpus que más lo
# necesitan).
OCR_ZOOM = 3.0


def _open_fitz_document(file_path: Path):
    """Abre el PDF con PyMuPDF para poder renderizar páginas a imagen (fallback OCR).

    Se abre una sola vez por documento (el llamador lo cachea) en vez de una vez
    por página, para no pagar el costo de reabrir el archivo repetidamente en
    documentos donde muchas páginas disparan el fallback.

    Devuelve `None` (en vez de levantar) si PyMuPDF no está instalado o el
    archivo no se puede abrir con él — quien llama simplemente conserva el
    texto que haya extraído `pypdf` para esas páginas en vez de abortar la
    ingesta del documento completo por la falta de una dependencia opcional de
    OCR.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.warning(
            "PyMuPDF ('pymupdf') no esta instalado; no se puede aplicar el fallback OCR a "
            "paginas con poco texto en '%s'.",
            file_path.name,
        )
        return None

    try:
        return fitz.open(str(file_path))
    except Exception as exc:  # noqa: BLE001 - cualquier fallo de apertura es no-fatal (es un fallback)
        logger.warning(
            "No se pudo abrir '%s' con PyMuPDF para el fallback OCR: %s", file_path.name, exc
        )
        return None


def _ocr_extract_page_text(fitz_doc, page_num: int, file_path: Path) -> str:
    """Renderiza la página `page_num` (1-based) del PDF ya abierto con PyMuPDF y
    le corre OCR (Tesseract, español) para recuperar texto que `pypdf` no pudo
    extraer (contenido renderizado como imagen, típicamente por DRM/marca de
    agua de biblioteca).

    No levanta excepción si algo falla: se loggea y se devuelve `""` — quien
    llama (`extract_pdf_pages`) conserva el texto original de `pypdf` para esa
    página en ese caso.
    """
    try:
        import fitz  # PyMuPDF
        import pytesseract
    except ImportError:
        logger.warning(
            "pymupdf/pytesseract no estan instalados; no se puede aplicar el fallback OCR a "
            "la pagina %s de '%s'.",
            page_num,
            file_path.name,
        )
        return ""

    try:
        page = fitz_doc[page_num - 1]
    except IndexError:
        logger.warning(
            "PyMuPDF no encontro la pagina %s en '%s' (fallback OCR).", page_num, file_path.name
        )
        return ""

    try:
        pixmap = page.get_pixmap(matrix=fitz.Matrix(OCR_ZOOM, OCR_ZOOM))
    except Exception as exc:  # noqa: BLE001 - fallback: cualquier fallo de render no es fatal
        logger.warning(
            "No se pudo renderizar la pagina %s de '%s' para OCR: %s",
            page_num,
            file_path.name,
            exc,
        )
        return ""

    tmp_path: str | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_file:
            tmp_path = tmp_file.name
        pixmap.save(tmp_path)
        return pytesseract.image_to_string(tmp_path, lang="spa")
    except Exception as exc:  # noqa: BLE001 - fallback: cualquier fallo de OCR no es fatal
        logger.warning("OCR fallo en la pagina %s de '%s': %s", page_num, file_path.name, exc)
        return ""
    finally:
        if tmp_path is not None:
            try:
                os.remove(tmp_path)
            except OSError:  # pragma: no cover - limpieza best-effort
                pass


class ExtractionError(ValueError):
    """El archivo tiene una extensión soportada pero no se pudo extraer su contenido.

    Cubre tanto "la librería de extracción no está instalada" como "el archivo no
    se pudo abrir" (corrupto, encriptado sin password, etc.).
    """


def extract_pdf_pages(file_path: Path) -> list[tuple[int, str]]:
    """Extrae texto por página de un PDF: `[(page_num_1_based, texto), ...]`.

    El número de página es el número de página PDF REAL (1-based), a diferencia
    de la convención de "índice secuencial de fragmento" que usa `chunk_text()`
    para `.md`/`.txt`/`.docx`/`.xlsx` (ver `app/rag/chunking.py`).

    Una página cuya extracción individual falla (o que simplemente no tiene texto
    embebido, p. ej. una página escaneada como imagen) se devuelve con texto vacío
    (`""`) en vez de abortar el documento completo — `app/rag/ingestion.py` filtra
    naturalmente las páginas vacías (`chunk_text("")` devuelve `[]`, sin chunks
    para esa página).

    Raises:
        ExtractionError: si el PDF completo no se puede abrir (corrupto, o
            encriptado con un password que no sea el vacío).
    """
    try:
        from pypdf import PdfReader
        from pypdf.errors import PdfReadError
    except ImportError as exc:  # pragma: no cover - depende de deps instaladas en Docker
        raise ExtractionError(
            "La librería 'pypdf' no está instalada; no se puede extraer texto de PDF."
        ) from exc

    try:
        reader = PdfReader(str(file_path))
    except (PdfReadError, OSError, ValueError) as exc:
        raise ExtractionError(f"No se pudo abrir el PDF '{file_path.name}': {exc}") from exc

    if reader.is_encrypted:
        try:
            decrypt_result = reader.decrypt("")
        except Exception as exc:  # pypdf lanza distintas excepciones según el backend de cifrado
            raise ExtractionError(
                f"PDF encriptado y no se pudo desencriptar sin password: "
                f"'{file_path.name}' ({exc})"
            ) from exc
        if not decrypt_result:
            raise ExtractionError(
                f"PDF encriptado con password requerido (no se pudo abrir con password "
                f"vacío): '{file_path.name}'"
            )

    pages: list[tuple[int, str]] = []
    total_pages = len(reader.pages)
    # Se abre lazy (solo si alguna pagina dispara el fallback) para no pagar el
    # costo de PyMuPDF en el ~99% de paginas del corpus que ya extraen bien con
    # pypdf. Una vez abierto se reusa para el resto de las paginas del mismo
    # documento en vez de reabrirlo por pagina.
    fitz_doc = None
    try:
        for page_num, page in enumerate(reader.pages, start=1):
            try:
                text = page.extract_text() or ""
            except Exception as exc:  # noqa: BLE001 - defensa amplia intencional: ver docstring del módulo
                logger.warning(
                    "No se pudo extraer texto de la pagina %s de '%s': %s. Se indexa como vacia.",
                    page_num,
                    file_path.name,
                    exc,
                )
                text = ""

            stripped = text.strip()
            if len(stripped) < OCR_MIN_CHARS:
                if fitz_doc is None:
                    fitz_doc = _open_fitz_document(file_path)
                if fitz_doc is not None:
                    print(
                        f"[extract_pdf_pages] Fallback OCR: pagina {page_num}/{total_pages} de "
                        f"'{file_path.name}' (pypdf extrajo {len(stripped)} caracteres, "
                        f"debajo del umbral de {OCR_MIN_CHARS})."
                    )
                    ocr_text = _ocr_extract_page_text(fitz_doc, page_num, file_path)
                    if len(ocr_text.strip()) > len(stripped):
                        text = ocr_text

            if not text.strip():
                logger.info(
                    "Pagina %s de '%s' no produjo texto (posible pagina escaneada/imagen).",
                    page_num,
                    file_path.name,
                )
            pages.append((page_num, text))
    finally:
        if fitz_doc is not None:
            fitz_doc.close()
    return pages


def extract_docx_text(file_path: Path) -> str:
    """Extrae y concatena el texto de todos los párrafos de un `.docx`.

    No hay concepto de "página" real recuperable sin renderizar el documento (Word
    no guarda saltos de página como metadata fiable) — no se inventa una. Quien
    llama a esta función (`app/rag/ingestion.py`) usa el índice secuencial de
    chunk como `page`, misma convención que `.md`/`.txt`.

    Raises:
        ExtractionError: si el archivo no se puede abrir (corrupto, no es
            realmente un .docx válido, etc.).
    """
    try:
        from docx import Document
    except ImportError as exc:  # pragma: no cover - depende de deps instaladas en Docker
        raise ExtractionError(
            "La librería 'python-docx' no está instalada; no se puede extraer texto de DOCX."
        ) from exc

    try:
        document = Document(str(file_path))
    except Exception as exc:  # noqa: BLE001 - cualquier fallo de apertura es fatal para este archivo
        raise ExtractionError(f"No se pudo abrir el DOCX '{file_path.name}': {exc}") from exc

    paragraphs = [p.text for p in document.paragraphs if p.text and p.text.strip()]
    return "\n".join(paragraphs)


def extract_xlsx_text(file_path: Path) -> str:
    """Representa un `.xlsx` como texto plano: por hoja, nombre de hoja + filas.

    Tampoco hay "página" real (una hoja puede tener cualquier cantidad de filas) —
    quien llama a esta función usa el índice secuencial de chunk como `page`,
    misma convención que `.md`/`.txt`/`.docx`.

    Raises:
        ExtractionError: si el archivo no se puede abrir (corrupto, no es
            realmente un .xlsx válido, etc.).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depende de deps instaladas en Docker
        raise ExtractionError(
            "La librería 'openpyxl' no está instalada; no se puede extraer texto de XLSX."
        ) from exc

    try:
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - cualquier fallo de apertura es fatal para este archivo
        raise ExtractionError(f"No se pudo abrir el XLSX '{file_path.name}': {exc}") from exc

    lines: list[str] = []
    try:
        for sheet in workbook.worksheets:
            lines.append(f"# Hoja: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(v) for v in row if v is not None]
                if values:
                    lines.append(" | ".join(values))
    finally:
        workbook.close()

    return "\n".join(lines)
